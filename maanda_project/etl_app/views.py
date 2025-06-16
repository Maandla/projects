
from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from openpyxl import Workbook
from .models import StudentScore
from io import BytesIO
from django.contrib import messages
from django.shortcuts import redirect
def index(request):
    return HttpResponse("Welcome to the ETL App!")



def upload_file(request):
    print("hello")
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        errors = []
        duplicates = []
        created_count = 0
        headers = [
            "gender", "race_ethnicity", "parental_level_of_education",
            "lunch", "test_preparation_course",
            "math_score", "reading_score", "writing_score"
        ]

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                gender = row[0]
                race_ethnicity = row[1]
                parental_level_of_education = row[2]
                lunch = row[3]
                test_preparation_course = row[4]

                # Validate score fields
                math_score = int(row[5])
                reading_score = int(row[6])
                writing_score = int(row[7])

                # Check for duplicates
                if StudentScore.objects.filter(
                    gender=gender,
                    race_ethnicity=race_ethnicity,
                    parental_level_of_education=parental_level_of_education,
                    lunch=lunch,
                    test_preparation_course=test_preparation_course,
                    math_score=math_score,
                    reading_score=reading_score,
                    writing_score=writing_score,
                ).exists():
                    duplicates.append(row)
                else:
                    StudentScore.objects.create(
                        gender=gender,
                        race_ethnicity=race_ethnicity,
                        parental_level_of_education=parental_level_of_education,
                        lunch=lunch,
                        test_preparation_course=test_preparation_course,
                        math_score=math_score,
                        reading_score=reading_score,
                        writing_score=writing_score,
                    )
                    created_count += 1

            except (ValueError, TypeError) as e:
                errors.append((row_index, row, str(e)))
        if errors:
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.title = "Invalid Rows"
            error_ws.append(headers + ["Error", "Excel Row"])

            for row_index, row_data, error_message in errors:
                error_ws.append(list(row_data) + [error_message, row_index])

            output = BytesIO()
            error_wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=upload_errors.xlsx'
            return response
            

        if duplicates:
            messages.warning(request, f"{len(duplicates)} duplicates found. {created_count} new rows uploaded.")
            return render(request, 'upload.html')
        messages.success(request, "Excel data uploaded successfully!")
        return render(request, 'upload.html')

    return render(request, 'upload.html')






def student_view(request):
    scores = StudentScore.objects.all()
    return render(request, 'display.html', {'scores': scores})



