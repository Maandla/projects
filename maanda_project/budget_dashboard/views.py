# views.py
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Sum, F, Q
from django.contrib import messages
from datetime import datetime
from decimal import Decimal
from collections import defaultdict, deque
from .models import Department, Project, BudgetLine, Expenditure
import openpyxl

def _quarter_from_month(m: int) -> int:
    return (m - 1) // 3 + 1


def _next_quarter(year: int, quarter: int):
    """Return (year, quarter) for the next quarter."""
    if quarter == 4:
        return year + 1, 1
    return year, quarter + 1


def dashboard_view(request):
    # -------- Filters (GET) --------
    year = request.GET.get('year')  # e.g., "2025" or ""
    quarter = request.GET.get('quarter')  # "1".."4" or ""
    department = request.GET.get('department')  # department name or ""

    # -------- Allocated (group by dept/year/quarter) --------
    allocated_qs = (
        BudgetLine.objects
        .values('project__department__name', 'year', 'quarter')
        .annotate(total_allocated=Sum('allocated'))
    )
    if year:
        allocated_qs = allocated_qs.filter(year=int(year))
    if quarter:
        allocated_qs = allocated_qs.filter(quarter=int(quarter))
    if department:
        allocated_qs = allocated_qs.filter(project__department__name=department)

    # -------- Expenditures (filter at Python level by year/quarter/department) --------
    expenditures = Expenditure.objects.select_related('project__department')
    if year or quarter or department:
        filtered = []
        Yr = int(year) if year else None
        Qr = int(quarter) if quarter else None
        for exp in expenditures:
            exp_year = exp.date.year
            exp_quarter = _quarter_from_month(exp.date.month)
            if ((Yr is None or exp_year == Yr) and
                (Qr is None or exp_quarter == Qr) and
                (not department or exp.project.department.name == department)):
                filtered.append(exp)
        expenditures = filtered
    else:
        expenditures = list(expenditures)

    # -------- Spent map: (dept, y, q) -> amount --------
    spent_map = defaultdict(Decimal)
    for exp in expenditures:
        y = exp.date.year
        q = _quarter_from_month(exp.date.month)
        key = (exp.project.department.name, y, q)
        spent_map[key] += exp.amount

    # -------- Dashboard rows (Allocated vs Spent per dept/quarter) --------
    dashboard_data = []
    for alloc in allocated_qs:
        dept = alloc['project__department__name']
        y = alloc['year']
        q = alloc['quarter']
        allocated = alloc['total_allocated'] or Decimal('0')
        spent = spent_map.get((dept, y, q), Decimal('0'))
        remaining = allocated - spent
        dashboard_data.append({
            'department': dept,
            'year': y,
            'quarter': q,
            'allocated': float(allocated),
            'spent': float(spent),
            'remaining': float(remaining),
        })
    dashboard_data.sort(key=lambda x: (x['year'], x['quarter'], x['department']))

    # -------- Totals --------
    total_allocated = float(sum(d['allocated'] for d in dashboard_data))
    total_spent = float(sum(d['spent'] for d in dashboard_data))
    total_remaining = float(total_allocated - total_spent)

    # =========================
    # Extra series for other charts
    # =========================

    # 1) Yearly trend (Allocated vs Spent per year) — not filter-bound; shows global trend
    yearly_allocated_qs = BudgetLine.objects.values('year').annotate(total_allocated=Sum('allocated'))
    yearly_spent_map = defaultdict(Decimal)
    for exp in Expenditure.objects.all():
        yearly_spent_map[exp.date.year] += exp.amount
    yearly_chart = []
    for row in yearly_allocated_qs:
        yr = row['year']
        yearly_chart.append({
            'year': yr,
            'allocated': float(row['total_allocated'] or 0),
            'spent': float(yearly_spent_map.get(yr, 0)),
        })
    yearly_chart.sort(key=lambda x: x['year'])

    # 2) Pie: Spent by Department (based on current filtered dashboard_data)
    dept_totals = defaultdict(lambda: {'allocated': 0.0, 'spent': 0.0})
    for row in dashboard_data:
        dept_totals[row['department']]['allocated'] += row['allocated']
        dept_totals[row['department']]['spent'] += row['spent']
    pie_chart = [{'name': dept, 'y': round(vals['spent'], 2)} for dept, vals in dept_totals.items()]

    # =========================
    # Forecasting (simple baseline)
    # =========================
    # Goal: show a forecast chart per department (next quarter forecast vs next quarter allocation if present)
    # Approach: For each department, take *historical quarterly spend* across all years in chronological order
    # and predict next quarter as the mean of the last up-to-4 quarters (or mean of all available if < 4).
    # This is simple/transparent and gives a baseline risk signal (overspend risk if forecast > allocation).

    # Build full historical spent by department by (year, quarter) across all data
    hist_spent_by_dept = defaultdict(lambda: defaultdict(Decimal))
    all_exps = Expenditure.objects.select_related('project__department').all()
    for exp in all_exps:
        d = exp.project.department.name
        y, q = exp.date.year, _quarter_from_month(exp.date.month)
        hist_spent_by_dept[d][(y, q)] += exp.amount

    # Determine reference period for "next quarter"
    # If the user selected a year/quarter, forecast the quarter *after* that.
    # Otherwise, forecast after the latest year/quarter that exists in data.
    if year and quarter:
        ref_year = int(year)
        ref_quarter = int(quarter)
    else:
        # find max (year, quarter) across *all* BudgetLine and Expenditure
        max_year, max_quarter = None, None
        # Consider allocations
        for bl in BudgetLine.objects.values_list('year', 'quarter'):
            y, q = bl
            if (max_year is None) or (y, q) > (max_year, max_quarter):
                max_year, max_quarter = y, q
        # Consider expenditures
        for exp in Expenditure.objects.values_list('date', flat=False):
            # exp is a date object via values_list('date', flat=False)
            # To keep things simple, re-query properly:
            pass
        # Better: iterate objects (we already did above)
        for exp in Expenditure.objects.all():
            y, q = exp.date.year, _quarter_from_month(exp.date.month)
            if (max_year is None) or (y, q) > (max_year, max_quarter):
                max_year, max_quarter = y, q
        ref_year, ref_quarter = max_year or 0, max_quarter or 1

    next_y, next_q = _next_quarter(ref_year, ref_quarter)

    # Compute forecast per department
    forecast_chart = []  # list of {department, forecast_spent, allocation_next, overspend}
    # To respect department filter in the forecast chart, limit departments accordingly
    dept_names = (
        [department] if department
        else list(Department.objects.order_by('name').values_list('name', flat=True).distinct())
    )

    for dept in dept_names:
        # historical spend time series for this department
        hist_items = sorted(hist_spent_by_dept[dept].items(), key=lambda t: t[0])  # [((y,q), amount), ...]
        last_vals = deque(maxlen=4)
        for (_, _), amt in hist_items:
            last_vals.append(float(amt))
        if len(last_vals) == 0:
            forecast_spent = 0.0
        else:
            forecast_spent = round(sum(last_vals) / len(last_vals), 2)

        # find next quarter allocation, if exists
        next_alloc_qs = (
            BudgetLine.objects
            .filter(project__department__name=dept, year=next_y, quarter=next_q)
            .aggregate(total_allocated=Sum('allocated'))
        )
        allocation_next = float(next_alloc_qs['total_allocated'] or 0.0)
        overspend = forecast_spent > allocation_next if allocation_next > 0 else False

        forecast_chart.append({
            'department': dept,
            'forecast_spent': forecast_spent,
            'allocation_next': allocation_next,
            'overspend': overspend,
            'next_year': next_y,
            'next_quarter': next_q,
        })

    # -------- Filter lists --------
    years = BudgetLine.objects.order_by('year').values_list('year', flat=True).distinct()
    quarters = [1, 2, 3, 4]
    departments = Department.objects.order_by('name').values_list('name', flat=True).distinct()

    # -------- AJAX response --------
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'dashboard_data': dashboard_data,
            'total_allocated': round(total_allocated, 2),
            'total_spent': round(total_spent, 2),
            'total_remaining': round(total_remaining, 2),
            'yearly_chart': yearly_chart,
            'pie_chart': pie_chart,
            'forecast_chart': forecast_chart,
            'forecast_meta': {
                'ref_year': ref_year,
                'ref_quarter': ref_quarter,
                'next_year': next_y,
                'next_quarter': next_q,
            }
        })

    # -------- Initial page render --------
    context = {
        'dashboard_data': dashboard_data,
        'total_allocated': total_allocated,
        'total_spent': total_spent,
        'total_remaining': total_remaining,

        'years': years,
        'quarters': quarters,
        'departments': departments,

        'selected_year': year or '',
        'selected_quarter': quarter or '',
        'selected_department': department or '',
    }
    return render(request, 'dashboard.html', context)

def dashboard_summary(request):
    current_year = datetime.now().year

    total_budget = BudgetLine.objects.filter(year=current_year).aggregate(total=Sum('allocated'))['total'] or 0
    total_spent = Expenditure.objects.filter(date__year=current_year).aggregate(total=Sum('amount'))['total'] or 0
    remaining = total_budget - total_spent

    data = {
        'year': current_year,
        'total_budget': float(total_budget),
        'total_spent': float(total_spent),
        'remaining': float(remaining)
    }
    return JsonResponse(data)

def project_quarterly_summary(request, project_id):
    year = request.GET.get("year")
    budget_qs = BudgetLine.objects.filter(project_id=project_id)
    expenditure_qs = Expenditure.objects.filter(project_id=project_id)

    if year:
        budget_qs = budget_qs.filter(year=year)
        expenditure_qs = expenditure_qs.filter(date__year=year)

    # Aggregate per quarter
    summary = []
    for q in range(1, 5):
        allocated = budget_qs.filter(quarter=q).aggregate(total=Sum("allocated"))["total"] or 0
        spent = expenditure_qs.filter(date__quarter=q).aggregate(total=Sum("amount"))["total"] or 0

        summary.append({
            "quarter": q,
            "allocated": float(allocated),
            "spent": float(spent),
            "balance": float(allocated - spent),
        })

    return JsonResponse({"project_id": project_id, "year": year, "quarters": summary})
from django.shortcuts import render

def dashboard_page(request):
    return render(request, 'dashboard.html')

from django.http import JsonResponse

def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please upload a file.")
            return JsonResponse({'redirect': '/dashboard/upload_excel/'}, status=200)

        try:
            wb = openpyxl.load_workbook(excel_file)

            for row in wb['Departments'].iter_rows(min_row=2, values_only=True):
                name = row[0]
                if name:
                    if Department.objects.filter(name=name).exists():
                        messages.error(request, f"Duplicate department found: {name}. Upload stopped.")
                        return JsonResponse({'redirect': '/dashboard/upload_excel/'}, status=200)
                    Department.objects.create(name=name)

            for row in wb['Projects'].iter_rows(min_row=2, values_only=True):
                dept_name, project_name = row[:2]
                if dept_name and project_name:
                    department = Department.objects.filter(name=dept_name).first()
                    if department:
                        Project.objects.create(department=department, name=project_name)

            for row in wb['BudgetLines'].iter_rows(min_row=2, values_only=True):
                project_name, year, quarter, allocated = row
                project = Project.objects.filter(name=project_name).first()
                if project and year and quarter and allocated is not None:
                    BudgetLine.objects.create(
                        project=project,
                        year=int(year),
                        quarter=int(quarter),
                        allocated=allocated
                    )

            for row in wb['Expenditures'].iter_rows(min_row=2, values_only=True):
                project_name, date, amount, description = row
                project = Project.objects.filter(name=project_name).first()
                if project and date and amount is not None:
                    Expenditure.objects.create(
                        project=project,
                        date=date,
                        amount=amount,
                        description=description
                    )

            messages.success(request, "Data imported successfully!")
            return JsonResponse({'redirect': '/dashboard/upload_excel/'}, status=200)

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return JsonResponse({'redirect': '/dashboard/upload_excel/'}, status=200)

    return render(request, 'upload_excel.html')



